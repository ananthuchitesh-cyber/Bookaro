#!/usr/bin/env python3
"""
Bookaro Travel Data Importer - All 32 Indian States
Reads xlsx files from state zip archives and loads into PostgreSQL travel_planner DB.

Usage:
    python scripts/import_travel_data.py
"""
import os, sys, re, zipfile, tempfile, shutil
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

DATABASE_URL = os.environ.get("DATABASE_URL", "postgresql://postgres:root@localhost:5432/travel_planner")
ZIP_FOLDER   = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "TravelApp"))

SCHEMA_STMTS = [
    """CREATE TABLE IF NOT EXISTS states (id SERIAL PRIMARY KEY, name TEXT NOT NULL UNIQUE, slug TEXT NOT NULL UNIQUE)""",
    """CREATE TABLE IF NOT EXISTS districts (
        id SERIAL PRIMARY KEY,
        state_id INTEGER NOT NULL REFERENCES states(id) ON DELETE CASCADE,
        name TEXT NOT NULL, slug TEXT NOT NULL,
        UNIQUE(state_id, slug))""",
    """CREATE TABLE IF NOT EXISTS tourist_places (
        id SERIAL PRIMARY KEY,
        district_id INTEGER NOT NULL REFERENCES districts(id) ON DELETE CASCADE,
        state_name TEXT NOT NULL, district_name TEXT NOT NULL, name TEXT NOT NULL,
        category TEXT, description TEXT, entry_fee TEXT, must_visit BOOLEAN DEFAULT FALSE,
        UNIQUE(district_id, name))""",
    """CREATE TABLE IF NOT EXISTS district_hotels (
        id SERIAL PRIMARY KEY,
        district_id INTEGER NOT NULL REFERENCES districts(id) ON DELETE CASCADE,
        state_name TEXT NOT NULL, district_name TEXT NOT NULL, hotel_name TEXT NOT NULL,
        category TEXT, star_rating NUMERIC(2,1), area TEXT,
        price_min INTEGER, price_max INTEGER, room_types TEXT, amenities TEXT,
        restaurant TEXT, ac_heating TEXT, parking TEXT, wifi TEXT, best_for TEXT, book_via TEXT,
        UNIQUE(district_id, hotel_name))""",
    """CREATE TABLE IF NOT EXISTS district_transport_routes (
        id SERIAL PRIMARY KEY,
        source_district_id INTEGER NOT NULL REFERENCES districts(id) ON DELETE CASCADE,
        dest_district_id   INTEGER NOT NULL REFERENCES districts(id) ON DELETE CASCADE,
        state_name TEXT NOT NULL, source_district TEXT NOT NULL, dest_district TEXT NOT NULL,
        distance_km INTEGER, bus_time TEXT, bus_ordinary_fare TEXT, bus_ac_fare TEXT,
        taxi_fare TEXT, taxi_time TEXT, train_available TEXT,
        train_station_from TEXT, train_station_to TEXT, train_fare TEXT, train_time TEXT,
        UNIQUE(source_district_id, dest_district_id))""",
    """CREATE TABLE IF NOT EXISTS district_travel_costs (
        id SERIAL PRIMARY KEY,
        source_district_id INTEGER NOT NULL REFERENCES districts(id) ON DELETE CASCADE,
        dest_district_id   INTEGER NOT NULL REFERENCES districts(id) ON DELETE CASCADE,
        state_name TEXT NOT NULL, source_district TEXT NOT NULL, dest_district TEXT NOT NULL,
        tier TEXT NOT NULL CHECK (tier IN ('budget','comfort','luxury')),
        travel_cost INTEGER, stay_per_day INTEGER,
        day1 INTEGER, day2 INTEGER, day3 INTEGER, day4 INTEGER, day5 INTEGER,
        day6 INTEGER, day7 INTEGER, day8 INTEGER, day9 INTEGER, day10 INTEGER,
        UNIQUE(source_district_id, dest_district_id, tier))""",
    "CREATE INDEX IF NOT EXISTS idx_tp_district ON tourist_places(district_id)",
    "CREATE INDEX IF NOT EXISTS idx_dh_district ON district_hotels(district_id)",
    "CREATE INDEX IF NOT EXISTS idx_dr_src_dst  ON district_transport_routes(source_district_id, dest_district_id)",
    "CREATE INDEX IF NOT EXISTS idx_dc_src_dst  ON district_travel_costs(source_district_id, dest_district_id, tier)",
    "CREATE INDEX IF NOT EXISTS idx_tp_state    ON tourist_places(state_name)",
    "CREATE INDEX IF NOT EXISTS idx_dh_state    ON district_hotels(state_name)",
]

# ── Helpers ──────────────────────────────────────────────────────────────────

def clean_int(val):
    if val is None: return None
    s = str(val).replace(',','').replace('\u20b9','').replace('INR','').strip()
    m = re.search(r'\d+', s)
    return int(m.group()) if m else None

def clean_text(val, maxlen=None):
    if val is None: return None
    s = re.sub(r'^[^\w\u0900-\u097f]+', '', str(val)).strip()
    return (s[:maxlen] if maxlen else s) or None

def slug(name):
    return re.sub(r'[^a-z0-9]+', '_', name.lower()).strip('_')

def parse_stars(val):
    if val is None: return None
    m = re.search(r'(\d+\.?\d*)', str(val))
    if not m: return None
    return min(float(m.group()), 9.9)

def parse_price_range(val):
    if val is None: return None, None
    nums = re.findall(r'\d+', str(val).replace(',',''))
    if len(nums) >= 2: return int(nums[0]), int(nums[1])
    if len(nums) == 1: return int(nums[0]), int(nums[0])
    return None, None

def cost_val(val):
    if val is None: return None
    s = str(val).strip()
    if s in ('\u2014','\u2013','-','','N/A','\u2014'): return None
    return clean_int(s)

SKIP_SHEETS = {'Sheet','HOW TO USE','Distance Matrix','Bus Fare Matrix',
               'Train Fare Matrix','GUIDE','FULL MATRIX - Budget Day1'}

# ── DB helpers ───────────────────────────────────────────────────────────────

def upsert_state(cur, name):
    s = slug(name)
    cur.execute("INSERT INTO states(name,slug) VALUES(%s,%s) ON CONFLICT(slug) DO UPDATE SET name=EXCLUDED.name RETURNING id", (name,s))
    return cur.fetchone()[0]

def upsert_district(cur, state_id, name):
    s = slug(name)
    cur.execute("INSERT INTO districts(state_id,name,slug) VALUES(%s,%s,%s) ON CONFLICT(state_id,slug) DO UPDATE SET name=EXCLUDED.name RETURNING id", (state_id,name,s))
    return cur.fetchone()[0]

# ── File parsers ─────────────────────────────────────────────────────────────

def import_itinerary(wb, cur, state_id, state_name):
    places = []
    for sh in wb.sheetnames:
        if sh in SKIP_SHEETS: continue
        dist = sh.strip()
        did = upsert_district(cur, state_id, dist)
        ws = wb[sh]
        reading = False
        ni,ci,di,ei,mi = 1,2,3,4,5
        for row in ws.iter_rows(values_only=True):
            if not row: continue
            if str(row[0] or '').strip() == '#':
                hdr = [str(c or '').strip().lower() for c in row]
                ni = next((i for i,h in enumerate(hdr) if 'place' in h or 'name' in h),1)
                ci = next((i for i,h in enumerate(hdr) if 'category' in h),2)
                di = next((i for i,h in enumerate(hdr) if 'desc' in h),3)
                ei = next((i for i,h in enumerate(hdr) if 'entry' in h or 'fee' in h),4)
                mi = next((i for i,h in enumerate(hdr) if 'must' in h),5)
                reading = True; continue
            if not reading: continue
            if str(row[0] or '').strip().isdigit():
                g = lambda i: row[i] if i < len(row) else None
                nm = clean_text(g(ni))
                if nm:
                    mv = '\u2b50' in str(g(mi) or '') or 'yes' in str(g(mi) or '').lower()
                    places.append((did,state_name,dist,nm,clean_text(g(ci)),clean_text(g(di),500),clean_text(g(ei)),mv))
    if places:
        seen = {}
        for r in places: seen[(r[0],r[3])] = r
        places = list(seen.values())
        execute_values(cur,
            "INSERT INTO tourist_places(district_id,state_name,district_name,name,category,description,entry_fee,must_visit) VALUES %s "
            "ON CONFLICT(district_id,name) DO UPDATE SET category=EXCLUDED.category,description=EXCLUDED.description,entry_fee=EXCLUDED.entry_fee,must_visit=EXCLUDED.must_visit",
            places)
    return len(places)

def import_hotels(wb, cur, state_id, state_name):
    hotels = []
    for sh in wb.sheetnames:
        if sh in SKIP_SHEETS: continue
        dist = sh.strip()
        did = upsert_district(cur, state_id, dist)
        ws = wb[sh]
        reading = False
        ni,cati,stri,ai,pi,ri,ameni,resti,aci,parki,wii,bfi,bki = 1,2,3,4,5,6,7,8,9,10,11,12,13
        for row in ws.iter_rows(values_only=True):
            if not row: continue
            if str(row[0] or '').strip() == '#':
                hdr = [str(c or '').strip().lower() for c in row]
                def h(kw,d): return next((i for i,x in enumerate(hdr) if kw in x),d)
                ni=h('hotel name',1); cati=h('category',2); stri=h('star',3)
                ai=h('location',4); pi=h('price',5); ri=h('room',6)
                ameni=h('amenities',7); resti=h('restaurant',8); aci=h('ac',9)
                parki=h('parking',10); wii=h('wi-fi',11); wii=h('wifi',wii)
                bfi=h('best for',12); bki=h('book',13)
                reading=True; continue
            if not reading: continue
            if str(row[0] or '').strip().isdigit():
                g = lambda i: row[i] if i < len(row) else None
                nm = clean_text(g(ni))
                if not nm: continue
                cat = clean_text(g(cati))
                if cat:
                    cl = cat.lower()
                    cat = 'budget' if 'budget' in cl else ('luxury' if 'luxury' in cl else 'comfort')
                pmn,pmx = parse_price_range(g(pi))
                hotels.append((did,state_name,dist,nm,cat,parse_stars(g(stri)),clean_text(g(ai)),
                               pmn,pmx,clean_text(g(ri),200),clean_text(g(ameni),300),
                               clean_text(g(resti),100),clean_text(g(aci),50),
                               clean_text(g(parki),50),clean_text(g(wii),50),
                               clean_text(g(bfi),100),clean_text(g(bki),200)))
    if hotels:
        execute_values(cur,
            "INSERT INTO district_hotels(district_id,state_name,district_name,hotel_name,category,star_rating,area,price_min,price_max,room_types,amenities,restaurant,ac_heating,parking,wifi,best_for,book_via) VALUES %s "
            "ON CONFLICT(district_id,hotel_name) DO UPDATE SET category=EXCLUDED.category,star_rating=EXCLUDED.star_rating,area=EXCLUDED.area,price_min=EXCLUDED.price_min,price_max=EXCLUDED.price_max,room_types=EXCLUDED.room_types,amenities=EXCLUDED.amenities,restaurant=EXCLUDED.restaurant,ac_heating=EXCLUDED.ac_heating,parking=EXCLUDED.parking,wifi=EXCLUDED.wifi,best_for=EXCLUDED.best_for,book_via=EXCLUDED.book_via",
            hotels)
    return len(hotels)

def import_transport(wb, cur, state_id, state_name):
    routes = []
    for sh in wb.sheetnames:
        if sh in SKIP_SHEETS: continue
        src = sh.strip()
        sid = upsert_district(cur, state_id, src)
        ws = wb[sh]
        reading = False
        for row in ws.iter_rows(values_only=True):
            if not row: continue
            if any('to district' in str(v or '').lower() for v in row[:3]):
                reading=True; continue
            if not reading: continue
            dest_raw = clean_text(row[0] if len(row)>0 else None)
            if not dest_raw or 'to district' in dest_raw.lower(): continue
            dest_name = re.sub(r'^[^\w]+','',dest_raw).strip()
            if not dest_name: continue
            g = lambda i: row[i] if i < len(row) else None
            did = upsert_district(cur, state_id, dest_name)
            routes.append((sid,did,state_name,src,dest_name,
                           clean_int(g(1)),clean_text(g(2),50),clean_text(g(3),100),clean_text(g(4),100),
                           clean_text(g(5),100),clean_text(g(6),50),clean_text(g(7),100),
                           clean_text(g(8),200),clean_text(g(9),200),clean_text(g(10),100),clean_text(g(11),50)))
    if routes:
        execute_values(cur,
            "INSERT INTO district_transport_routes(source_district_id,dest_district_id,state_name,source_district,dest_district,distance_km,bus_time,bus_ordinary_fare,bus_ac_fare,taxi_fare,taxi_time,train_available,train_station_from,train_station_to,train_fare,train_time) VALUES %s "
            "ON CONFLICT(source_district_id,dest_district_id) DO UPDATE SET distance_km=EXCLUDED.distance_km,bus_time=EXCLUDED.bus_time,bus_ordinary_fare=EXCLUDED.bus_ordinary_fare,bus_ac_fare=EXCLUDED.bus_ac_fare,taxi_fare=EXCLUDED.taxi_fare,taxi_time=EXCLUDED.taxi_time,train_available=EXCLUDED.train_available,train_station_from=EXCLUDED.train_station_from,train_station_to=EXCLUDED.train_station_to,train_fare=EXCLUDED.train_fare,train_time=EXCLUDED.train_time",
            routes)
    return len(routes)

def import_costs(wb, cur, state_id, state_name):
    """
    Per-district sheet = destination. Each row = source district.
    Header row: #, From District, Dist(km), BUDGET\nTravel, BUDGET\nStay/Day, BUDGET\nD1..D10,
                COMFORT\n..., LUXURY\n...
    """
    costs = []
    for sh in wb.sheetnames:
        if sh in SKIP_SHEETS: continue
        if 'FULL MATRIX' in sh.upper(): continue
        dest = sh.strip()
        dest_id = upsert_district(cur, state_id, dest)
        ws = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        data_start = None
        tier_offsets = {}      # tier -> column index of "Travel" cell
        for ri, row in enumerate(rows):
            vals = [str(v or '').strip().lower() for v in row]
            # Header: has '#' and 'from district' or similar
            if '#' in vals and any('district' in v or 'destination' in v or 'from' in v for v in vals if v):
                data_start = ri + 1
                # Scan this header row for BUDGET/COMFORT/LUXURY keywords
                for ci, v in enumerate(row):
                    sv = str(v or '').upper()
                    if 'BUDGET' in sv and 'budget' not in tier_offsets:
                        tier_offsets['budget'] = ci
                    elif 'COMFORT' in sv and 'comfort' not in tier_offsets:
                        tier_offsets['comfort'] = ci
                    elif 'LUXURY' in sv and 'luxury' not in tier_offsets:
                        tier_offsets['luxury'] = ci
                break
        if data_start is None or not tier_offsets:
            continue
        for row in rows[data_start:]:
            if not row or row[0] is None: continue
            if not str(row[0]).strip().isdigit(): continue
            src_raw = str(row[1] if len(row) > 1 else '').strip()
            src_name = re.sub(r'^[^\w]+', '', src_raw).strip()
            if not src_name: continue
            src_id = upsert_district(cur, state_id, src_name)
            for tier, off in tier_offsets.items():
                g = lambda d, o=off: row[o+d] if (o+d) < len(row) else None
                travel = cost_val(g(0))
                stay   = cost_val(g(1))
                # D1,D2,D3,D5,D7,D10 = up to 8 values → pad to 10
                raw = [cost_val(g(2+i)) for i in range(8)]
                day_vals = (raw + [None, None])[:10]
                costs.append((src_id, dest_id, state_name, src_name, dest, tier, travel, stay, *day_vals))
    if costs:
        seen = {}
        for row in costs: seen[(row[0],row[1],row[5])] = row
        costs = list(seen.values())
        execute_values(cur,
            "INSERT INTO district_travel_costs(source_district_id,dest_district_id,state_name,source_district,dest_district,tier,travel_cost,stay_per_day,day1,day2,day3,day4,day5,day6,day7,day8,day9,day10) VALUES %s "
            "ON CONFLICT(source_district_id,dest_district_id,tier) DO UPDATE SET travel_cost=EXCLUDED.travel_cost,stay_per_day=EXCLUDED.stay_per_day,day1=EXCLUDED.day1,day2=EXCLUDED.day2,day3=EXCLUDED.day3,day4=EXCLUDED.day4,day5=EXCLUDED.day5,day6=EXCLUDED.day6,day7=EXCLUDED.day7,day8=EXCLUDED.day8,day9=EXCLUDED.day9,day10=EXCLUDED.day10",
            costs)
    return len(costs)

# ── State name map ───────────────────────────────────────────────────────────
STATE_MAP = {
    'andhra':'Andhra Pradesh','arunachalpradesh':'Arunachal Pradesh','assam':'Assam',
    'bihar':'Bihar','chandigarh':'Chandigarh','chattisgarh':'Chhattisgarh',
    'delhi':'Delhi','goa':'Goa','gujarat':'Gujarat','haryana':'Haryana',
    'himalchal':'Himachal Pradesh','jamu':'Jammu & Kashmir','jharkhand':'Jharkhand',
    'karanatak':'Karnataka','kerela':'Kerala','ladakh':'Ladakh',
    'maharastha':'Maharashtra','manipur':'Manipur','meghalaya':'Meghalaya',
    'mizoram':'Mizoram','mp':'Madhya Pradesh','nagaland':'Nagaland',
    'odisha':'Odisha','punjab':'Punjab','rajathan':'Rajasthan','sikkim':'Sikkim',
    'tn':'Tamil Nadu','telangana':'Telangana','tripura':'Tripura',
    'up':'Uttar Pradesh','uttarkhand':'Uttarakhand','westbengal':'West Bengal',
}

def zip_state(zf):
    key = re.sub(r'[^a-z]', '', os.path.splitext(zf)[0].lower())
    for pat,nm in STATE_MAP.items():
        if key.startswith(pat) or key == pat: return nm
    return os.path.splitext(zf)[0].replace('_',' ').title()

def file_type(fn):
    f = fn.lower()
    if 'hotel' in f: return 'hotels'
    if 'transport_guide' in f: return 'transport'
    if 'all10days' in f or 'budget_10days' in f or 'area_to_area' in f or 'district_to_district' in f: return 'costs'
    if 'itinerary' in f: return 'itinerary'
    return None

def import_zip(zpath, conn, state_name):
    tmpdir = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(zpath,'r') as z: z.extractall(tmpdir)
        xlsxs = []
        for r,_,fs in os.walk(tmpdir):
            for f in fs:
                if f.endswith('.xlsx') and not f.startswith('~$'):
                    xlsxs.append(os.path.join(r,f))
        cur = conn.cursor()
        state_id = upsert_state(cur, state_name)
        conn.commit()
        totals = {k:0 for k in ('itinerary','hotels','transport','costs')}
        for xp in xlsxs:
            ft = file_type(os.path.basename(xp))
            if not ft: print(f"    skip: {os.path.basename(xp)}"); continue
            print(f"    -> {ft}: {os.path.basename(xp)}")
            try:
                wb = openpyxl.load_workbook(xp, read_only=True, data_only=True)
                if ft=='itinerary':  totals['itinerary']  += import_itinerary(wb,cur,state_id,state_name)
                elif ft=='hotels':   totals['hotels']     += import_hotels(wb,cur,state_id,state_name)
                elif ft=='transport':totals['transport']  += import_transport(wb,cur,state_id,state_name)
                elif ft=='costs':    totals['costs']      += import_costs(wb,cur,state_id,state_name)
                wb.close(); conn.commit()
            except Exception as e:
                conn.rollback()
                print(f"    ERROR in {os.path.basename(xp)}: {e}")
        cur.close()
        return totals
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

def main():
    print("Bookaro Travel Data Importer")
    print(f"DB  : {DATABASE_URL}")
    print(f"ZIPs: {ZIP_FOLDER}\n")
    if not os.path.exists(ZIP_FOLDER):
        print(f"ERROR: ZIP folder not found: {ZIP_FOLDER}"); sys.exit(1)

    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()
    print("Creating schema...")
    for stmt in SCHEMA_STMTS:
        cur.execute(stmt)
    conn.commit(); cur.close()
    print("Schema OK\n")

    zips = sorted(f for f in os.listdir(ZIP_FOLDER) if f.endswith('.zip'))
    print(f"Found {len(zips)} zip files\n")
    grand = {k:0 for k in ('itinerary','hotels','transport','costs')}
    errors = []
    for zf in zips:
        sn = zip_state(zf)
        print(f"[{sn}] ({zf})")
        try:
            t = import_zip(os.path.join(ZIP_FOLDER,zf), conn, sn)
            print(f"  Places:{t['itinerary']} Hotels:{t['hotels']} Routes:{t['transport']} Costs:{t['costs']}")
            for k in grand: grand[k] += t[k]
        except Exception as e:
            print(f"  FAILED: {e}"); errors.append((zf,str(e))); conn.rollback()

    conn.close()
    print(f"\n--- Import Complete ---")
    print(f"Tourist places  : {grand['itinerary']:,}")
    print(f"Hotels          : {grand['hotels']:,}")
    print(f"Transport routes: {grand['transport']:,}")
    print(f"Travel cost rows: {grand['costs']:,}")
    if errors:
        print(f"\nErrors ({len(errors)}):")
        for zf,e in errors: print(f"  {zf}: {e}")

if __name__ == '__main__':
    main()
